import openpyxl
import re
import qrcode
import os


# sheet_name = 'Sorcerer-Wizard'
spell_list = {"dancing lights", "resistance", "ray of frost", "mage hand", "read magic", "prestidigitation",
              "detect magic", "detect poison", "mage armor", "burning hands", "disguise self", "identify",
              "chastise", "fire breath", "flaming sphere", "color spray", "see invisibility",
              "fireball", "arcane mark", "alarm", "burning arc", "message", "mending", "comprehend languages",
              "silent image", "eagle's splendor", "diminish resistance"}

print("Loading workbook")
wb = openpyxl.load_workbook('spell_full.xlsx')
sh = wb[wb.sheetnames[0]]

spell_list = [s.lower() for s in list(spell_list)]
spell_list.sort()
found_idx = 0
spell_num = len(spell_list)

if not os.path.isdir('./output'):
    print("Output folder not found, create it")
    os.mkdir('./output')

if not os.path.isdir('./output/qr'):
    print("Output folder not found, create it")
    os.mkdir('./output/qr')

no_link_list = []
for row in sh.iter_rows():
    name = row[0].value
    lower_name = name.lower()
    if lower_name in spell_list[:]:
        cell_text = row[70].value
        try:
            link = re.search('http://.*",', cell_text).group()[:-2]
        except AttributeError:
            no_link_list.append(name)
            link_name = ''.join([c if c.isalnum() else '-' for c in lower_name])
            link = f"https://www.d20pfsrd.com/magic/all-spells/{lower_name[0]}/{link_name}"

        img = qrcode.make(link)
        img.save(f"./output/qr/{name}.png")
        found_idx = found_idx + 1

        print(f"{found_idx}/{spell_num} found")
        spell_list.remove(lower_name)
        if not spell_list:
            break

if spell_list:
    print(f"The following spells have not been found: {spell_list}")

if no_link_list:
    print(f"No link has been found for following spells, but generated automatically: {no_link_list}")
print("Done")
